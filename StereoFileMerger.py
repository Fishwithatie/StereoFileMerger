import tkinter as tk
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
import os
import sys
from tkinter import messagebox
from tkinter import filedialog
from tkinter import simpledialog
import tkinter


def main(folder_path):

    row_number_main = 1 
    main_file = Workbook()
    ws1 = main_file.active
    value_dict = {}
    value_dict_cava = {}
    main_file_create(main_file, ws1)
    cava_list = []
    probe_list = []

    for fn in os.listdir(folder_path):
        fn = folder_path + "\\" + fn
        if "xlsx" in fn:
            filter_files_by_probe(fn, cava_list, probe_list)

    for files in probe_list:
        open_file_probe(files, value_dict)
        copy_to_main_probe(value_dict, main_file, ws1, row_number_main)
        row_number_main = len(value_dict[0]) + row_number_main
        print (files)

    for files in cava_list:
        open_file_cava(files, value_dict_cava)
        copy_to_main_cava(value_dict_cava, main_file, ws1)
        print (files)


    namefile = filedialog.asksaveasfilename(defaultextension = ".xlsx")
    main_file.save(filename = namefile)

def select_folder():
    
    folder_path = filedialog.askdirectory()
    folder_path = folder_path.replace("/" , "\\")
    print (folder_path)
    if folder_path != None:     
        messagebox.showinfo("Please Wait...", "Merging stereology data from folder " + folder_path)
        main(folder_path)
    else:
        messagebox.showinfo("ERROR!", "Please select a valid folder")



def main_file_create(main_file, ws1):
    
    
    ws1.title = "Raw Data"

    ws1["A1"] = "Data File"
    ws1["B1"] = "Marker"
    ws1["C1"] = "Total Markers Counted"
    ws1["D1"] = "Markers Counted as Half Counts"
    ws1["E1"] = "Adjusted Marker Count"
    ws1["F1"] = "Number of sections"
    ws1["G1"] = "Region"
    ws1["H1"] = "Number of Sampling Sites"
    ws1["I1"] = "User Defined Mounted Thickness"
    ws1["J1"] = "Measured Defined Mounted Thickness"
    ws1["K1"] = "Estimated Length by Spaceballs (µm)"
    ws1["L1"] = "Coefficient of Error (Gundersen), m=0 [1]"
    ws1["M1"] = "Coefficient of Error (Gundersen), m=1 [1]"
    ws1["N1"] = "Probe Shape"
    ws1["O1"] = "Probe Volume (µm³)"
    ws1["P1"] = "Sampling Grid Area (XY)(µm²)"
    ws1["Q1"] = "Volume Corrected for OverProjection (µm³)"

    return(main_file)


def open_file_probe(importfile, value_dict):
    
    importedfile = load_workbook(filename = importfile)
    summary = importedfile.active

    col_number = 0
    for col in summary.columns:
        value_dict[col_number] = []
        col_number = col_number + 1
        for cell in col:
            value_dict[col_number-1].append(cell.value)

    for i in value_dict:
            del value_dict[i][0]

    return (value_dict)

def open_file_cava(importfile, value_dict_cava):

    importedfile = load_workbook(filename = importfile)
    summary = importedfile.active

    for cell1 in summary.columns[1]:
        if cell1.value != "Marker":
            if isinstance(cell1.value, str): 
                cell2 = cell1.value.lower()
                value_dict_cava[cell2] = 0
                for volume in summary.columns[4]:
                    if volume.value == 0:
                        if volume.row == cell1.row:
                            value_dict_cava[cell2] = summary.cell(row = volume.row, column = 3).value
                    elif volume.row == cell1.row:
                        value_dict_cava[cell2] = (volume.value)

    return(value_dict_cava)


def filter_files_by_probe(fileID, cava_list, probe_list):

    importedfile = load_workbook(filename = fileID)
    summary = importedfile.active

    if summary["D1"].value == "Estimated Volume (µm³)":
        cava_list.append(fileID)
    else:
        probe_list.append(fileID)
    return(cava_list, probe_list)



def copy_to_main_probe(value_dict, main_file, ws1, row_number_main):
    
    for col in value_dict:
        col_number = col
        row_number = row_number_main
        for x in value_dict[col]:
            ws1.cell(row = row_number+1, column = col_number+1).value = x
            row_number = row_number + 1
            
            
    return (main_file,row_number)

def copy_to_main_cava(value_dict_cava, main_file, ws1):

    for i in value_dict_cava:
        for cell in ws1.columns[6]:
                if cell.value.lower() == i:                         
                    ws1.cell(row = cell.row, column = 17).value = value_dict_cava[i]
    return (main_file)


top = tkinter.Tk()

folderbutton = tk.Button(top, text = "Please select folder", command = select_folder)
folderbutton.pack(side="left")

EXIT = tk.Button(top, text="EXIT", command=sys.exit)
EXIT.pack(side="bottom")

top.title("Stereology Merger")
folderbutton.pack()
EXIT.pack()
top.mainloop()
